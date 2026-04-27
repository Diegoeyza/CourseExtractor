import os
import sys
import zipfile
import shutil
import tempfile
import threading
import tkinter as tk
from tkinter import filedialog, messagebox

import customtkinter as ctk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from extractor_service import CourseExtractor

# ── Theme ──────────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ── Colour palette (matches the web app's blue/emerald/gray palette) ───────────
CLR_HEADER_COURSE   = "1E3A5F"   # deep navy  → course header block
CLR_HEADER_PREREQ   = "374151"   # charcoal   → prerequisites section header
CLR_HEADER_RA       = "1E40AF"   # blue       → learning outcomes header
CLR_HEADER_APE      = "065F46"   # emerald    → profile contributions header
CLR_HEADER_BIBLIO   = "4B1D6E"   # purple     → bibliography header
CLR_SUBHEADER       = "F3F4F6"   # light gray → column-name rows
CLR_ROW_ALT         = "EFF6FF"   # very light blue → alternating data rows

THIN = Side(border_style="thin", color="D1D5DB")


# ── Excel helpers ──────────────────────────────────────────────────────────────

def _style_section_header(cell, hex_color: str, label: str):
    cell.value = label
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=hex_color)
    cell.alignment = Alignment(wrap_text=True, vertical="center")


def _style_col_header(cell, label: str):
    cell.value = label
    cell.font = Font(bold=True, color="111827", size=10)
    cell.fill = PatternFill("solid", fgColor=CLR_SUBHEADER)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = Border(bottom=Side(border_style="medium", color="9CA3AF"))


def _style_data(cell, value, alt_row=False):
    cell.value = value
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if alt_row:
        cell.fill = PatternFill("solid", fgColor=CLR_ROW_ALT)
    cell.border = Border(
        left=THIN, right=THIN, top=THIN, bottom=THIN
    )


def _merge_section_header(ws, row, hex_color, label):
    ws.merge_cells(f"A{row}:B{row}")
    _style_section_header(ws[f"A{row}"], hex_color, label)
    ws.row_dimensions[row].height = 22


def _blank_row(ws, row):
    ws.row_dimensions[row].height = 6


def write_course_sheet(wb: Workbook, data: dict):
    """Write a single course to a new worksheet in wb."""
    course  = data["course"]
    prereqs = data["prerequisites"]
    ras     = data["ras"]
    apes    = data["apes"]
    biblio  = data["bibliography"]

    # Sheet name: full_nrc or fallback to course name truncated
    sheet_name = course.get("full_nrc") or course.get("name", "Course")
    # Excel sheet names max 31 chars, no special chars
    sheet_name = sheet_name[:31].replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "").replace("[", "").replace("]", "").replace(":", "-")
    ws = wb.create_sheet(title=sheet_name)

    # Set column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80

    r = 1  # current row cursor

    # ── Course header block ───────────────────────────────────────────────────
    _merge_section_header(ws, r, CLR_HEADER_COURSE, "COURSE INFORMATION")
    r += 1

    meta_fields = [
        ("Name",     course.get("name", "")),
        ("Area",     course.get("area", "")),
        ("Code",     course.get("code", "")),
        ("Full NRC", course.get("full_nrc", "")),
    ]
    for label, value in meta_fields:
        ws[f"A{r}"].value = label
        ws[f"A{r}"].font = Font(bold=True, size=10)
        ws[f"A{r}"].fill = PatternFill("solid", fgColor="DBEAFE")
        ws[f"A{r}"].alignment = Alignment(vertical="center")
        ws[f"B{r}"].value = value
        ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 18
        r += 1

    # Description (merged, taller row)
    ws[f"A{r}"].value = "Description"
    ws[f"A{r}"].font = Font(bold=True, size=10)
    ws[f"A{r}"].fill = PatternFill("solid", fgColor="DBEAFE")
    ws[f"A{r}"].alignment = Alignment(vertical="top")
    ws[f"B{r}"].value = course.get("description", "")
    ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 80
    r += 1

    _blank_row(ws, r); r += 1

    # ── Prerequisites ─────────────────────────────────────────────────────────
    _merge_section_header(ws, r, CLR_HEADER_PREREQ, f"PRE-REQUISITE NETWORK  ({len(prereqs)} items)")
    r += 1
    _style_col_header(ws[f"A{r}"], "Name")
    _style_col_header(ws[f"B{r}"], "NRC")
    r += 1
    if prereqs:
        for i, p in enumerate(prereqs):
            alt = (i % 2 == 1)
            _style_data(ws[f"A{r}"], p.get("name", ""), alt)
            _style_data(ws[f"B{r}"], p.get("nrc", ""), alt)
            ws.row_dimensions[r].height = 16
            r += 1
    else:
        ws.merge_cells(f"A{r}:B{r}")
        ws[f"A{r}"].value = "No prerequisites listed"
        ws[f"A{r}"].font = Font(italic=True, color="6B7280")
        ws.row_dimensions[r].height = 16
        r += 1

    _blank_row(ws, r); r += 1

    # ── Learning Outcomes (RAs) ───────────────────────────────────────────────
    _merge_section_header(ws, r, CLR_HEADER_RA, f"LEARNING OUTCOMES — RA  ({len(ras)} items)")
    r += 1
    _style_col_header(ws[f"A{r}"], "ID")
    _style_col_header(ws[f"B{r}"], "Description")
    r += 1
    if ras:
        for i, ra in enumerate(ras):
            alt = (i % 2 == 1)
            _style_data(ws[f"A{r}"], ra.get("id", ""), alt)
            _style_data(ws[f"B{r}"], ra.get("description", ""), alt)
            ws.row_dimensions[r].height = 40
            r += 1
    else:
        ws.merge_cells(f"A{r}:B{r}")
        ws[f"A{r}"].value = "No learning outcomes found"
        ws[f"A{r}"].font = Font(italic=True, color="6B7280")
        ws.row_dimensions[r].height = 16
        r += 1

    _blank_row(ws, r); r += 1

    # ── Profile Contributions (APEs) ──────────────────────────────────────────
    _merge_section_header(ws, r, CLR_HEADER_APE, f"GRADUATION PROFILE CONTRIBUTIONS — APE  ({len(apes)} items)")
    r += 1
    _style_col_header(ws[f"A{r}"], "ID")
    _style_col_header(ws[f"B{r}"], "Description")
    r += 1
    if apes:
        for i, ape in enumerate(apes):
            alt = (i % 2 == 1)
            _style_data(ws[f"A{r}"], ape.get("id", ""), alt)
            _style_data(ws[f"B{r}"], ape.get("description", ""), alt)
            ws.row_dimensions[r].height = 40
            r += 1
    else:
        ws.merge_cells(f"A{r}:B{r}")
        ws[f"A{r}"].value = "No profile contributions found"
        ws[f"A{r}"].font = Font(italic=True, color="6B7280")
        ws.row_dimensions[r].height = 16
        r += 1

    _blank_row(ws, r); r += 1

    # ── Basic Bibliography ────────────────────────────────────────────────────
    _merge_section_header(ws, r, CLR_HEADER_BIBLIO, f"BASIC BIBLIOGRAPHY  ({len(biblio)} items)")
    r += 1
    _style_col_header(ws[f"A{r}"], "Title")
    _style_col_header(ws[f"B{r}"], "Metadata (ISBN / Author / Editor / Date)")
    r += 1
    if biblio:
        for i, item in enumerate(biblio):
            alt = (i % 2 == 1)
            _style_data(ws[f"A{r}"], item.get("title", ""), alt)
            _style_data(ws[f"B{r}"], item.get("metadata", ""), alt)
            ws.row_dimensions[r].height = 30
            r += 1
    else:
        ws.merge_cells(f"A{r}:B{r}")
        ws[f"A{r}"].value = "No bibliography listed"
        ws[f"A{r}"].font = Font(italic=True, color="6B7280")
        ws.row_dimensions[r].height = 16
        r += 1


def build_excel(pdf_paths: list, output_path: str, progress_cb=None):
    """
    Extract data from each PDF and write one sheet per course to output_path.
    progress_cb(current, total, message) if provided.
    Returns (success_count, errors) where errors is list of (filename, msg).
    """
    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    errors = []
    total = len(pdf_paths)

    for i, pdf_path in enumerate(pdf_paths):
        fname = os.path.basename(pdf_path)
        if progress_cb:
            progress_cb(i, total, f"Processing {fname}…")
        try:
            extractor = CourseExtractor(pdf_path)
            data = extractor.get_structured_data()
            write_course_sheet(wb, data)
        except Exception as e:
            errors.append((fname, str(e)))

    if progress_cb:
        progress_cb(total, total, "Saving Excel file…")

    wb.save(output_path)
    return total - len(errors), errors


# ── GUI ────────────────────────────────────────────────────────────────────────

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Course Extractor")
        self.geometry("780x560")
        self.minsize(640, 480)
        self.resizable(True, True)

        # State
        self._files: list[str] = []   # flat list of resolved PDF paths
        self._source_dirs: list[str] = []  # dirs of original uploads (for default output)
        self._running = False

        self._build_ui()

    # ── UI construction ────────────────────────────────────────────────────────

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        # ── Top bar ────────────────────────────────────────────────────────────
        top = ctk.CTkFrame(self, height=52, corner_radius=0)
        top.grid(row=0, column=0, sticky="ew")
        top.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(top, text="Course Extractor", font=ctk.CTkFont(size=17, weight="bold")).grid(
            row=0, column=0, padx=18, pady=12, sticky="w"
        )
        ctk.CTkLabel(top, text="PDF → Excel", font=ctk.CTkFont(size=12), text_color="gray60").grid(
            row=0, column=1, padx=0, pady=12, sticky="w"
        )

        # ── Drop / browse area ─────────────────────────────────────────────────
        drop_frame = ctk.CTkFrame(self, corner_radius=10, border_width=2, border_color="#3B82F6")
        drop_frame.grid(row=1, column=0, padx=20, pady=(14, 6), sticky="nsew")
        drop_frame.grid_columnconfigure(0, weight=1)
        drop_frame.grid_rowconfigure(2, weight=1)

        # Header of drop zone (Label + Buttons)
        drop_header = ctk.CTkFrame(drop_frame, fg_color="transparent")
        drop_header.grid(row=0, column=0, padx=12, pady=(10, 0), sticky="ew")
        drop_header.grid_columnconfigure(0, weight=1)

        # Drop-zone label
        hint = ctk.CTkLabel(
            drop_header,
            text="Add PDF files or a ZIP archive containing PDFs",
            font=ctk.CTkFont(size=13),
            text_color="gray60",
            anchor="w"
        )
        hint.grid(row=0, column=0, sticky="w", padx=4)

        btn_row = ctk.CTkFrame(drop_header, fg_color="transparent")
        btn_row.grid(row=0, column=1, sticky="e")
        ctk.CTkButton(btn_row, text="Add PDFs", width=110, command=self._browse_pdfs).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="Add ZIP", width=110, command=self._browse_zip, fg_color="#6D28D9", hover_color="#5B21B6").pack(side="left")
        ctk.CTkButton(btn_row, text="Clear", width=80, fg_color="#374151", hover_color="#4B5563", command=self._clear_files).pack(side="left", padx=(8, 0))

        # File list
        self._listbox_frame = ctk.CTkScrollableFrame(drop_frame, label_text="", corner_radius=6)
        self._listbox_frame.grid(row=2, column=0, padx=12, pady=12, sticky="nsew")
        self._listbox_frame.grid_columnconfigure(0, weight=1)

        self._file_labels: list[ctk.CTkLabel] = []
        self._empty_label = ctk.CTkLabel(
            self._listbox_frame,
            text="No files selected yet",
            text_color="gray50",
            font=ctk.CTkFont(size=12, slant="italic"),
        )
        self._empty_label.grid(row=0, column=0, padx=8, pady=8)

        # ── Status / progress ─────────────────────────────────────────────────
        status_frame = ctk.CTkFrame(self, fg_color="transparent")
        status_frame.grid(row=2, column=0, padx=20, pady=2, sticky="ew")
        status_frame.grid_columnconfigure(0, weight=1)

        self._status_label = ctk.CTkLabel(
            status_frame, text="", font=ctk.CTkFont(size=11), text_color="gray60", anchor="w"
        )
        self._status_label.grid(row=0, column=0, sticky="ew")

        self._progress = ctk.CTkProgressBar(status_frame, height=6)
        self._progress.grid(row=1, column=0, sticky="ew", pady=(2, 0))
        self._progress.set(0)
        self._progress.grid_remove()

        # ── Bottom action bar ─────────────────────────────────────────────────
        bottom = ctk.CTkFrame(self, height=58, corner_radius=0)
        bottom.grid(row=3, column=0, sticky="ew")
        bottom.grid_columnconfigure(0, weight=1)

        self._output_info = ctk.CTkLabel(
            bottom,
            text="Output: will be saved next to the first input file",
            font=ctk.CTkFont(size=11),
            text_color="gray55",
            anchor="w",
        )
        self._output_info.grid(row=0, column=0, padx=18, pady=0, sticky="w")

        btn_inner = ctk.CTkFrame(bottom, fg_color="transparent")
        btn_inner.grid(row=0, column=1, padx=14, pady=8)

        self._extract_btn = ctk.CTkButton(
            btn_inner,
            text="Extract → Excel",
            width=160,
            height=36,
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._start_extraction,
        )
        self._extract_btn.pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            btn_inner,
            text="Change Output",
            width=130,
            height=36,
            fg_color="#374151",
            hover_color="#4B5563",
            command=self._pick_output,
        ).pack(side="left")

        # Custom output path (None = auto)
        self._custom_output: str | None = None

    # ── File management ────────────────────────────────────────────────────────

    def _browse_pdfs(self):
        paths = filedialog.askopenfilenames(
            title="Select PDF files",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if paths:
            self._add_files(list(paths))

    def _browse_zip(self):
        path = filedialog.askopenfilename(
            title="Select ZIP archive",
            filetypes=[("ZIP archives", "*.zip"), ("All files", "*.*")],
        )
        if path:
            self._add_files([path])

    def _add_files(self, paths: list[str]):
        """Resolve ZIPs to their contained PDFs and add to the list."""
        temp_dirs_to_track = []
        for path in paths:
            self._source_dirs.append(os.path.dirname(os.path.abspath(path)))
            if path.lower().endswith(".zip"):
                try:
                    tmp = tempfile.mkdtemp(prefix="course_extractor_")
                    temp_dirs_to_track.append(tmp)
                    with zipfile.ZipFile(path, "r") as zf:
                        zf.extractall(tmp)
                    for root, _, files in os.walk(tmp):
                        for f in files:
                            if f.lower().endswith(".pdf"):
                                self._files.append(os.path.join(root, f))
                except Exception as e:
                    messagebox.showerror("ZIP Error", f"Could not open ZIP:\n{e}")
            elif path.lower().endswith(".pdf"):
                self._files.append(path)

        # Deduplicate while preserving order
        seen = set()
        deduped = []
        for f in self._files:
            if f not in seen:
                seen.add(f)
                deduped.append(f)
        self._files = deduped

        self._refresh_file_list()
        self._update_output_info()

    def _clear_files(self):
        self._files.clear()
        self._source_dirs.clear()
        self._custom_output = None
        self._refresh_file_list()
        self._update_output_info()

    def _refresh_file_list(self):
        for lbl in self._file_labels:
            lbl.destroy()
        self._file_labels.clear()

        if not self._files:
            self._empty_label.grid()
        else:
            self._empty_label.grid_remove()
            for i, path in enumerate(self._files):
                fname = os.path.basename(path)
                lbl = ctk.CTkLabel(
                    self._listbox_frame,
                    text=f"  {fname}",
                    font=ctk.CTkFont(size=11),
                    anchor="w",
                    text_color="#93C5FD" if i % 2 == 0 else "#6EE7B7",
                )
                lbl.grid(row=i, column=0, sticky="ew", pady=1)
                self._file_labels.append(lbl)

    def _update_output_info(self):
        if self._custom_output:
            self._output_info.configure(text=f"Output: {self._custom_output}")
        elif self._files and self._source_dirs:
            folder = self._source_dirs[0]
            self._output_info.configure(
                text=f"Output: {folder}/courses.xlsx  (auto — next to first input)"
            )
        else:
            self._output_info.configure(text="Output: will be saved next to the first input file")

    def _pick_output(self):
        path = filedialog.asksaveasfilename(
            title="Save Excel file as…",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="courses.xlsx",
        )
        if path:
            self._custom_output = path
            self._update_output_info()

    # ── Extraction ─────────────────────────────────────────────────────────────

    def _resolve_output_path(self) -> str:
        if self._custom_output:
            return self._custom_output
        folder = self._source_dirs[0] if self._source_dirs else os.path.expanduser("~")
        return os.path.join(folder, "courses.xlsx")

    def _start_extraction(self):
        if self._running:
            return
        if not self._files:
            messagebox.showwarning("No files", "Please add at least one PDF or ZIP file.")
            return

        self._running = True
        self._extract_btn.configure(state="disabled", text="Processing…")
        self._progress.grid()
        self._progress.set(0)
        self._set_status("Starting…")

        output_path = self._resolve_output_path()

        def worker():
            def cb(current, total, msg):
                frac = current / total if total else 0
                self.after(0, lambda: self._progress.set(frac))
                self.after(0, lambda: self._set_status(msg))

            success, errors = build_excel(self._files, output_path, progress_cb=cb)

            def finish():
                self._running = False
                self._extract_btn.configure(state="normal", text="Extract → Excel")
                self._progress.set(1)

                if errors:
                    err_detail = "\n".join(f"• {f}: {m}" for f, m in errors)
                    if success > 0:
                        msg = f"{success} course(s) extracted successfully.\n\nErrors ({len(errors)}):\n{err_detail}"
                        messagebox.showwarning("Partial Success", msg)
                    else:
                        messagebox.showerror("Extraction Failed", f"All files failed:\n{err_detail}")
                    self._set_status(f"Done with errors — {success} succeeded, {len(errors)} failed")
                else:
                    self._set_status(f"Done! {success} course(s) saved to {output_path}")
                    messagebox.showinfo(
                        "Success",
                        f"{success} course(s) extracted.\n\nSaved to:\n{output_path}",
                    )

            self.after(0, finish)

        threading.Thread(target=worker, daemon=True).start()

    def _set_status(self, msg: str):
        self._status_label.configure(text=msg)


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.mainloop()
